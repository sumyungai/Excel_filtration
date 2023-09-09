# -*- coding: utf-8 -*-
"""
Created on Fri Sep  8 18:26:44 2023

@author: ruste
"""

import openpyxl

from openpyxl import Workbook

path = r"D:\Users\ruste\Documents\activeFile.xlsx"

data = openpyxl.load_workbook(path)


sheets = data.sheetnames

#specify the amount of sheets here
sheetOne = data[sheets[0]]

sheetTwo = data[sheets[1]]


ColNamesTwo = {}
Current  = 1
for COL in sheetTwo.iter_cols(1, sheetTwo.max_column):
    ColNamesTwo[Current] = COL[0].value
    Current += 1


ColNamesOne = {}
Current  = 1
for COL in sheetOne.iter_cols(1, sheetOne.max_column):
    ColNamesOne[Current] = COL[0].value
    Current += 1


Final_data = Workbook()
for row in sheetTwo.iter_cols(min_row=2, min_col=6, max_row = sheetTwo.max_row, max_col= sheetTwo.max_column):
    for cell in row:
        if (cell.value not in Final_data.sheetnames):
            Final_data.create_sheet(cell.value)
            
            
#Specify general row structure                  
rowShape = ["Poster Number", "First Name", "Last Name", "Abstract Title"]
for sheet in Final_data:
        sheet.append(rowShape)
        
        
data_group_b = []
#Specify columns you want to include
judge_names = ["First Name","Judge 1", "Judge 2", "Judge 3", "Judge 4"]
for row in sheetTwo.iter_rows(min_row=2, min_col=1, max_row= sheetTwo.max_row, max_col=sheetTwo.max_column):
    temp_data = []
    i =0
    for cell in row:
        if (ColNamesTwo[cell.column] in judge_names):
            temp_data.append(cell.value)
            i+=1
    data_group_b.append(temp_data)
  
   
data_group_a = []
#Specify columns you want to include
poster_names = ["First Name","Last Name","Display Stands Numbers", "Abstract Title"]
for row in sheetOne.iter_rows(min_row=2, min_col=1, max_row=sheetOne.max_row, max_col=sheetOne.max_column):
    temp_data = []
    i = 0
    for cell in row:
        if (ColNamesOne[cell.column] in poster_names):
            temp_data.append(cell.value)
            i+=1
    data_group_a.append(temp_data)


for person in data_group_a:
    for lst in data_group_b:
        
        if (person[1].lower().strip() == lst[0].lower().strip()):
            for i in range(1,5):Final_data[lst[-i]].append(person) 
#Specify the name you want to save the file as            
Final_data.save('JudgeData.xlsx')